#!/usr/bin/env python3
"""
script for scraping product pages from Amazon, BestBuy, and Samsung using Playwright.
- Saves each page's HTML to a unique file in outputs/ directory.
- Parses each HTML to extract product price and model number (or SKU for Samsung).

This version additionally saves the run's result as a single row in an Excel file:
- Columns = keys (flattened per-site/per-index keys like "amazon_1_url")
- Row = values for this run
- On next run the script appends a new row (does not overwrite previous data).

"""
from zoneinfo import ZoneInfo
import datetime
import asyncio
import json
import os
import random
import re

from urllib.parse import quote_plus
from playwright.async_api import async_playwright, TimeoutError
from bs4 import BeautifulSoup
from openpyxl.utils import column_index_from_string, get_column_letter
# new imports for Excel writing
from openpyxl import Workbook, load_workbook

# =========================================================================
# script_2.py fixes (vs script.py)
# -------------------------------------------------------------------------
# 1. Redirect detection: many product pages now silently redirect to a
#    *different* product (e.g. S26 Ultra -> S25 FE, Z Fold 7 -> Z Flip 7) when
#    the item is out of stock/unavailable. We compare the product identifier we
#    REQUESTED (ASIN / BestBuy code / Samsung SKU) against the identifier in the
#    page's <link rel="canonical">. On mismatch we record "not available".
# 2. Duplicate-element fix: the old hardcoded selectors matched many elements on
#    the page (related items, sponsored, carousels), causing wrong reads. We now
#    scope price extraction to the MAIN price container only.
# 3. Updated selectors for current UI: Amazon price -> corePriceDisplay /
#    priceToPay; BestBuy & Samsung price -> JSON-LD offer (Samsung keyed by SKU).
# 4. BestBuy S26 fix: those pages fail Playwright's HTTP/2 with
#    ERR_HTTP2_PROTOCOL_ERROR, so nothing saved. We launch BestBuy's Chromium
#    with --disable-http2 (forces HTTP/1.1) so S26 pages load.
# 5. results.xlsx now uses the same layout as "Price Comparisons_v3_WIP":
#    51 product groups x 9 columns starting at column C, timestamp in column B.
# Everything else (URLs, delays, user agents, cookies logic) is unchanged.
# =========================================================================
NOT_AVAILABLE = "not available"

# Excel layout of Price Comparisons_v3_WIP (per product group of 9 columns):
#   +0 Amazon price  +1 Samsung price  +2 BestBuy price
#   +3 SKU_Amazon    +4 SKU_Samsung    +5 SKU_BestBuy
#   +6 vs Amazon (formula, left blank)  +7 vs Bestbuy (formula, left blank)  +8 blank
FIRST_GROUP_COL = 3        # column C
GROUP_STRIDE    = 9
TIMESTAMP_COL   = 2        # column B

# Product group headers for row 1 (matches the WIP workbook, 51 slots in order)
PRODUCT_LABELS = [
    "Z Fold 7 512 GB", "S25 Ultra 512GB", "Z Flip 512 GB", "S25 Edge 512 GB",
    "Galaxy Tab S11 512 GB", "Z Fold 7 256GB Blue Shadow", "Z Fold 7 256GB Jet Black",
    "Z Fold 7 256GB Silver Shadow", "Z Fold 7 1TB Zet Black", "Z Fold 7 512GB Zet Black",
    "Z Fold 7 512GB Silver Shadow", "S25 Ultra 256GB Titanium Black",
    "S25 Ultra 256GB Titanium Gray", "S25 Ultra 256GB Titanium Silver Blue",
    "S25 Ultra 256GB Titanium White Silver", "S25 Ultra 512GB Titanium Gray",
    "S25 Ultra 512GB Titanium Silver Blue", "S25 Ultra 512GB Titanium White Silver",
    "S25 Edge 256GB Titanium Silver", "S25 Edge 512GB Titanium Jet Black",
    "S26 Ultra 256GB Sky Blue", "S26 Ultra 256GB Cobalt Violet", "S26 Ultra 256GB White",
    "S26 Ultra 256GB Black", "S26 Ultra 512GB Sky Blue", "S26 Ultra 512GB Cobalt Violet",
    "S26 Ultra 512GB White", "S26 Ultra 512GB Black", "S26 Ultra 1TB Sky Blue ",
    "S26 Ultra 1TB Cobalt Violet", "S26 Ultra 1TB White", "S26 Ultra 1TB Black",
    "S26 256GB Sky Blue", "S26 256GB Cobalt Violet", "S26 256GB White", "S26 256GB Black",
    "S26 512GB Sky Blue", "S26 512GB Cobalt Violet", "S26 512GB White", "S26 512GB Black",
    "S26+ 256GB Sky Blue", "S26+ 256GB Cobalt Violet", "S26+ 256GB White", "S26+ 256GB Black",
    "S26+ 512GB Sky Blue", "S26+ 512GB Cobalt Violet", "S26+ 512GB White", "S26+ 512GB Black",
]
SUBHEADERS = ["Amazon price", "Samsung price", "BestBuy.com price",
              "SKU_ID_Amazon ", "SKU_ID_Samsung ", "SKU_ID_BestBuy.com",
              "vs Amazon", "vs Bestbuy"]


# ---- product identifiers (used for redirect detection & slot SKUs) ----
def amazon_id_from_url(url):
    m = re.search(r"/dp/([A-Z0-9]{10})", url or "", re.I)
    return m.group(1).upper() if m else None

def bestbuy_id_from_url(url):
    m = re.search(r"/product/[^/]+/([A-Z0-9]+)", url or "", re.I)
    return m.group(1).upper() if m else None

def get_canonical_href(html):
    """Pull <link rel=canonical href=...> without a full DOM parse."""
    m = re.search(r'<link\b[^>]*\brel=["\']canonical["\'][^>]*>', html, re.I)
    if not m:
        return None
    h = re.search(r'href=["\']([^"\']+)["\']', m.group(0), re.I)
    return h.group(1) if h else None

def iter_ldjson(html):
    """Yield parsed JSON-LD objects from the HTML (regex-sliced, fast)."""
    for m in re.finditer(
            r'<script\b[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.I | re.S):
        try:
            data = json.loads(m.group(1).strip())
        except Exception:
            continue
        for it in (data if isinstance(data, list) else [data]):
            yield it


# ---- price cleaning (ported from ConvertDirtyTextPriceToNumbers/main2) ----
def clean_price_value(raw):
    """Return a float rounded to 2dp, or None. Mirrors main2_decimalPlaceTill2."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "" or s == NOT_AVAILABLE:
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    s = re.sub(r"[^\d\.,\-]", "", s)
    if s == "" or re.fullmatch(r"[-\.,]*", s):
        return None
    s = s.replace("−", "-")
    if "-" in s:
        if s.count("-") > 1:
            s = s.replace("-", "")
        if s.startswith("-"):
            negative = not negative
            s = s.lstrip("-")
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
            if s.count(".") > 1:
                left, right = s.rsplit(".", 1)
                s = left.replace(".", "") + "." + right
    elif has_comma and not has_dot:
        parts = s.split(",")
        if len(parts) >= 2 and len(parts[-1]) == 2:
            s = ",".join(parts[:-1]).replace(",", "") + "." + parts[-1]
        else:
            s = s.replace(",", "")
    elif has_dot and not has_comma:
        if s.count(".") > 1:
            left, right = s.rsplit(".", 1)
            s = left.replace(".", "") + "." + right
    s = re.sub(r"[^\d.]", "", s)
    if s.count(".") > 1:
        left, right = s.rsplit(".", 1)
        s = left.replace(".", "") + "." + right
    if s in ("", "."):
        return None
    try:
        value = round(float(s), 2)
    except Exception:
        return None
    return -value if negative else value


# -----------------------
# Shared helpers
# -----------------------
async def human_delay(min_sec=0.5, max_sec=2.5):
    """Wait for a random time between min_sec and max_sec seconds."""
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)

async def human_delay_short():
    """Small helper to yield control briefly (kept minimal to respect original logic)."""
    await asyncio.sleep(0.1)

async def get_page_content_safe(page, retries=4):
    """Return page HTML, tolerating in-flight client-side navigations.

    BestBuy fires a delayed client-side navigation/reload ~20s after load, which
    made a bare `page.content()` throw:
      "Unable to retrieve content because the page is navigating and changing".
    We wait for the page to settle and retry; as a last resort we read
    document.documentElement.outerHTML via JS (works mid-navigation).
    """
    last_err = None
    for attempt in range(retries):
        try:
            # let any in-flight navigation finish before grabbing content
            try:
                await page.wait_for_load_state("domcontentloaded", timeout=8000)
            except Exception:
                pass
            return await page.content()
        except Exception as e:
            last_err = e
            # brief settle, then retry
            await asyncio.sleep(2.5)
    # final fallback: pull the DOM directly (succeeds even while navigating)
    try:
        return await page.evaluate("() => document.documentElement.outerHTML")
    except Exception:
        raise last_err

def sanitize_filename(s: str, maxlen: int = 200) -> str:
    """Create a filesystem-safe short filename from a string (URL)."""
    if not s:
        return "file"
    s_enc = quote_plus(s, safe="")
    s_clean = re.sub(r'[^A-Za-z0-9._-]', '_', s_enc)
    return s_clean[:maxlen]

def _to_jsonable(v):
    """Convert complex types to JSON strings for Excel storage; leave primitives as-is."""
    if isinstance(v, (dict, list, tuple)):
        return json.dumps(v, ensure_ascii=False)
    return v

def _render_cells(result, slot_sku, site):
    """Given a per-URL result dict, return (price_cell, sku_cell) for the sheet.

    - price -> float when parseable, "not available" for redirect/no-price, else None.
    - sku   -> canonical per-slot SM code (Amazon/BestBuy upper, Samsung lower);
               "not available" mirrors the price when the product wasn't found.
    """
    raw = result.get("price") if result else None
    if raw == NOT_AVAILABLE:
        return NOT_AVAILABLE, NOT_AVAILABLE
    num = clean_price_value(raw)
    if num is None:
        return None, None            # genuine gap (fetch error / empty URL) -> blank
    sku = None
    if slot_sku:
        sku = slot_sku if site == "samsung" else slot_sku.upper()
    return num, sku


def save_results_wip_format(am_res, bb_res, sam_res, samsung_urls, ts_str,
                            excel_path="outputs/results.xlsx"):
    """Append one row per run to results.xlsx using the SAME layout as
    'Price Comparisons_v3_WIP.xlsx':
      - row 1 = product group headers, row 2 = sub-headers, data from row 3
      - 51 groups x 9 columns starting at column C; timestamp in column B
      - each group: Amazon/Samsung/BestBuy price, 3 SKU columns, 2 'vs'
        formula columns (filled with the same formulas as the WIP file), 1 blank
    Prices are written as numbers; SKU columns filled; 'vs' formulas added per
    row. Existing rows are kept.
    """
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=2, column=TIMESTAMP_COL, value="Timestamp EST")
        for s in range(len(PRODUCT_LABELS)):
            gc = FIRST_GROUP_COL + GROUP_STRIDE * s
            ws.cell(row=1, column=gc, value=PRODUCT_LABELS[s])
            for off, name in enumerate(SUBHEADERS):
                ws.cell(row=2, column=gc + off, value=name)

    r = ws.max_row + 1 if ws.max_row >= 2 else 3
    ws.cell(row=r, column=TIMESTAMP_COL, value=ts_str)

    n = len(PRODUCT_LABELS)
    for s in range(n):
        gc = FIRST_GROUP_COL + GROUP_STRIDE * s
        slot_sku = extract_sku_from_url(samsung_urls[s]) if s < len(samsung_urls) else None
        slot_sku = slot_sku.lower() if slot_sku else None

        for off, (res_list, site) in enumerate([
                (am_res, "amazon"), (sam_res, "samsung"), (bb_res, "bestbuy")]):
            result = res_list[s] if s < len(res_list) else None
            price_cell, sku_cell = _render_cells(result, slot_sku, site)
            pc = ws.cell(row=r, column=gc + off, value=price_cell)
            if isinstance(price_cell, float):
                pc.number_format = "0.00"
            ws.cell(row=r, column=gc + 3 + off, value=sku_cell)

        # +6 vs Amazon, +7 vs Bestbuy : same formulas as the WIP file, added on
        # every data row so they're in place as rows accumulate.
        #   vs Amazon  = Amazon price  / Samsung price - 1
        #   vs Bestbuy = BestBuy price / Samsung price - 1
        amazon_col  = get_column_letter(gc + 0)
        samsung_col = get_column_letter(gc + 1)
        bestbuy_col = get_column_letter(gc + 2)
        ws.cell(row=r, column=gc + 6,
                value=f"={amazon_col}{r}/{samsung_col}{r}-1")
        ws.cell(row=r, column=gc + 7,
                value=f"={bestbuy_col}{r}/{samsung_col}{r}-1")
        # +8 blank separator : intentionally left untouched

    wb.save(excel_path)
    print(f"✅ Results appended (WIP layout) to {excel_path} at row {r}")


def save_dict_to_excel_row(data: dict, excel_path: str = "outputs/results.xlsx"):
    """
    Save the provided dict as a single row in an Excel file.
    - Keys become column headers (first row).
    - Values become the next available row.
    - If the file exists, new keys are appended as new columns; existing column order is preserved.
    """
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        headers = list(data.keys())
        ws.append(headers)
        row = [ _to_jsonable(data.get(h)) for h in headers ]
        ws.append(row)
        wb.save(excel_path)
        print(f"✅ Results written to new Excel file: {excel_path}")
        return

    # file exists - load and append
    wb = load_workbook(excel_path)
    ws = wb.active

    # read existing headers from first row
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    existing_headers = [cell.value for cell in first_row]

    # compute headers union preserving existing order and appending new keys at the end
    new_keys = [k for k in data.keys() if k not in existing_headers]
    if new_keys:
        headers = existing_headers + new_keys
        # rewrite header row with expanded headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        headers = existing_headers

    # build row in header order
    row = []
    for h in headers:
        v = data.get(h)
        row.append(_to_jsonable(v) if v is not None else None)

    ws.append(row)
    wb.save(excel_path)

    #making changes from here
    column_refs = [
        "blank","a","d","cf","ar","e","cg","as","blank","blank","blank","i","ck","aw","j","cl","ax","blank","blank","blank","n","cp","bb","o","cq","bc","blank","blank","blank","s","cu","bg","t","cv","bh","blank","blank","blank","x","cz","bl","y","da","bm","blank","blank","blank","ac","de","bq","ad","df","br","blank","blank","blank","ah","dj","bv","ai","dk","bw","blank","blank","blank","am","do","ca","an","dp","cb","blank","blank","blank","dt","gb","ex","du","gc","ey","blank","blank","blank","dy","gg","fc","dz","gh","fd","blank","blank","blank","ed","gl","fh","ee","gm","fi","blank","blank","blank","ei","gq","fm","ej","gr","fn","blank","blank","blank","en","gv","fr","eo","gw","fs","blank","blank","blank","es","ha","fw","et","hb","fx", 'blank', 'blank', 'blank', 'hf', 'kr', 'iy', 'hg', 'ks', 'iz', 'blank', 'blank', 'blank', 'hk', 'kw', 'jd', 'hl', 'kx', 'je', 'blank', 'blank', 'blank', 'hp', 'lb', 'ji', 'hq', 'lc', 'jj', 'blank', 'blank', 'blank', 'hu', 'lg', 'jn', 'hv', 'lh', 'jo', 'blank', 'blank', 'blank', 'hz', 'll', 'js', 'ia', 'lm', 'jt', 'blank', 'blank', 'blank', 'ie', 'lq', 'jx', 'if', 'lr', 'jy', 'blank', 'blank', 'blank', 'ij', 'lv', 'kc', 'ik', 'lw', 'kd', 'blank', 'blank', 'blank', 'io', 'ma', 'kh', 'ip', 'mb', 'ki', 'blank', 'blank', 'blank', 'it', 'mf', 'km', 'iu', 'mg', 'kn', 'blank', 'blank', 'blank', 'mk', 'xe', 'ru', 'ml', 'xf', 'rv', 'blank', 'blank', 'blank', 'mp', 'xj', 'rz', 'mq', 'xk', 'sa', 'blank', 'blank', 'blank', 'mu', 'xo', 'se', 'mv', 'xp', 'sf', 'blank', 'blank', 'blank', 'mz', 'xt', 'sj', 'na', 'xu', 'sk', 'blank', 'blank', 'blank', 'ne', 'xy', 'so', 'nf', 'xz', 'sp', 'blank', 'blank', 'blank', 'nj', 'yd', 'st', 'nk', 'ye', 'su', 'blank', 'blank', 'blank', 'no', 'yi', 'sy', 'np', 'yj', 'sz', 'blank', 'blank', 'blank', 'nt', 'yn', 'td', 'nu', 'yo', 'te', 'blank', 'blank', 'blank', 'ny', 'ys', 'ti', 'nz', 'yt', 'tj', 'blank', 'blank', 'blank', 'od', 'yx', 'tn', 'oe', 'yy', 'to', 'blank', 'blank', 'blank', 'oi', 'zc', 'ts', 'oj', 'zd', 'tt', 'blank', 'blank', 'blank', 'on', 'zh', 'tx', 'oo', 'zi', 'ty', 'blank', 'blank', 'blank', 'os', 'zm', 'uc', 'ot', 'zn', 'ud', 'blank', 'blank', 'blank', 'ox', 'zr', 'uh', 'oy', 'zs', 'ui', 'blank', 'blank', 'blank', 'pc', 'zw', 'um', 'pd', 'zx', 'un', 'blank', 'blank', 'blank', 'ph', 'aab', 'ur', 'pi', 'aac', 'us', 'blank', 'blank', 'blank', 'pm', 'aag', 'uw', 'pn', 'aah', 'ux', 'blank', 'blank', 'blank', 'pr', 'aal', 'vb', 'ps', 'aam', 'vc', 'blank', 'blank', 'blank', 'pw', 'aaq', 'vg', 'px', 'aar', 'vh', 'blank', 'blank', 'blank', 'qb', 'aav', 'vl', 'qc', 'aaw', 'vm', 'blank', 'blank', 'blank', 'qg', 'aba', 'vq', 'qh', 'abb', 'vr', 'blank', 'blank', 'blank', 'ql', 'abf', 'vv', 'qm', 'abg', 'vw', 'blank', 'blank', 'blank', 'qq', 'abk', 'wa', 'qr', 'abl', 'wb', 'blank', 'blank', 'blank', 'qv', 'abp', 'wf', 'qw', 'abq', 'wg', 'blank', 'blank', 'blank', 'ra', 'abu', 'wk', 'rb', 'abv', 'wl', 'blank', 'blank', 'blank', 'rf', 'abz', 'wp', 'rg', 'aca', 'wq', 'blank', 'blank', 'blank', 'rk', 'ace', 'wu', 'rl', 'acf', 'wv', 'blank', 'blank', 'blank', 'rp', 'acj', 'wz', 'rq', 'ack', 'xa',
    ]
    new_sheet_base_name="SelectedColumns"
    source_sheet_name=None
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    # new_name = new_sheet_base_name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    # tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1
    wb.save(excel_path)
    print(f"✅ Results appended to Excel file: {excel_path}")





def copy_columns_by_references(
    file_path: str,
    column_refs: list,
    source_sheet_name: str | None = None,
    new_sheet_base_name: str = "CopiedColumns"
) -> str:
    """
    Copy columns from source sheet to a new sheet using Excel column letters.
    - column_refs: list of strings, column letters like ['A','D','X','AR', ...] or 'blank column'
    - source_sheet_name: None -> first sheet is used
    Returns the name of the created sheet.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1

    # Save workbook (overwrites existing file)
    wb.save(file_path)
    return new_name


# -----------------------
# AMAZON-specific logic
# -----------------------
async def save_amazon_htmls(
    urls,
    output_dir="outputs",
    cookies_file="amazon_cookies.json",
    headless=True,
):
    """Loop over the list of URLs, save each HTML to a unique file, and update cookies once."""
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless, slow_mo=100)

        # Load existing cookies/session state if available
        if os.path.exists(cookies_file):
            print("🍪 Loading existing cookies/session...")
            context = await browser.new_context(storage_state=cookies_file)
        else:
            print("🆕 No cookies found, creating a new session...")
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 768},
            )

        try:
            results = []
            for idx, url in enumerate(urls, start=1):
                # empty slot (e.g. product not yet listed): keep the position so
                # results stay aligned with the product groups, but skip cleanly.
                if not url or not url.strip():
                    print(f"\n[Amazon {idx}/{len(urls)}] empty URL slot -> skipping")
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": "empty"})
                    continue
                try:
                    safe_name = sanitize_filename(url)[:120]
                    output_file = os.path.join(output_dir, f"amazon_{idx}_{safe_name}.html")

                    page = await context.new_page()
                    print(f"\n[Amazon {idx}/{len(urls)}] Navigating to {url} ...")
                    await page.goto(url, wait_until="load")
                    await asyncio.sleep(10)  # Extra wait to ensure dynamic content loads

                    # Wait randomly for page content to settle
                    await human_delay(3, 6)

                    # 🖱️ Simulate random human-like mouse movement
                    for _ in range(3):
                        x = random.randint(200, 800)
                        y = random.randint(200, 600)
                        await page.mouse.move(x, y, steps=random.randint(5, 15))
                        await human_delay(0.3, 1.5)

                    # 🖱️ Random scrolling
                    for _ in range(2):
                        scroll_y = random.randint(400, 1000)
                        await page.mouse.wheel(0, scroll_y)
                        await human_delay(1, 3)

                    # Extract HTML (resilient to any mid-load client-side navigation)
                    html_content = await get_page_content_safe(page)
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print(f"✅ HTML saved to {output_file}")

                    # parse and collect results (updated: redirect-aware, scoped price)
                    price, redirected = parse_amazon_html(output_file, expected_url=url)

                    await page.close()

                    if redirected:
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "redirect"})
                    elif not price:
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "no_price"})
                    else:
                        # Amazon no longer exposes the SM- model on the page; the
                        # writer fills SKU_Amazon from the known per-slot SM code.
                        results.append({"url": url, "file": output_file, "price": price, "model": None, "status": "ok"})
                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        await page.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"})

            # Save cookies/session state after all pages are processed
            storage_state = await context.storage_state()
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(storage_state, f, ensure_ascii=False, indent=4)
            print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

def parse_amazon_html(html_file_path="amazon.html", expected_url=None):
    """Return (price_text_or_None, redirected_bool).

    - Redirect: compare requested ASIN vs the page's canonical link.
    - Price: scoped to the MAIN price container (corePriceDisplay / priceToPay)
      so we don't pick up sponsored/related prices elsewhere on the page.
    """
    if not os.path.exists(html_file_path):
        print(f"Error: HTML file '{html_file_path}' not found.")
        return None, False

    with open(html_file_path, "r", encoding="utf-8", errors="ignore") as file:
        html_content = file.read()

    # -------- REDIRECT DETECTION --------
    expected_asin = amazon_id_from_url(expected_url) if expected_url else None
    canonical = get_canonical_href(html_content)
    if expected_asin and canonical:
        can_asin = amazon_id_from_url(canonical)
        if can_asin and can_asin != expected_asin:
            print(f"[REDIRECT] requested {expected_asin} but page is {can_asin} -> not available")
            return None, True

    soup = BeautifulSoup(html_content, "lxml")

    # -------- PRICE EXTRACTION (scoped to the main buybox price container) --------
    price = None
    core = (soup.find(id="corePriceDisplay_desktop_feature_div")
            or soup.find(id="corePrice_feature_div")
            or soup.find(id="apex_desktop"))
    if core:
        pt = (core.find(class_="priceToPay")
              or core.find(class_="apexPriceToPay")
              or core)
        price_whole = pt.find("span", {"class": "a-price-whole"})
        price_fraction = pt.find("span", {"class": "a-price-fraction"})
        if price_whole:
            whole = re.sub(r"[^\d,]", "", price_whole.get_text())
            frac = re.sub(r"[^\d]", "", price_fraction.get_text()) if price_fraction else "00"
            price = f"{whole}.{frac or '00'}"
        else:
            for off in pt.find_all("span", {"class": "a-offscreen"}):
                t = off.get_text(strip=True)
                if t:
                    price = t
                    break

    if price:
        print(f"The price of the product is: {price}")
    else:
        print("Price not found in the HTML file.")

    return price, False


# -----------------------
# BESTBUY-specific logic
# -----------------------
def parse_bestbuy_html(input_file="bestbuy.html", expected_url=None):

    # Load HTML file
    if not os.path.exists(input_file):
        print(f"Error: HTML file '{input_file}' not found.")
        return None, None, False

    with open(input_file, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()

    # -------- REDIRECT DETECTION --------
    expected_code = bestbuy_id_from_url(expected_url) if expected_url else None
    canonical = get_canonical_href(html)
    if expected_code and canonical:
        can_code = bestbuy_id_from_url(canonical)
        if can_code and can_code != expected_code:
            print(f"[REDIRECT] requested {expected_code} but page is {can_code} -> not available")
            return None, None, True

    # -------- PRICE + MODEL from JSON-LD (single, reliable source) --------
    price = None
    model_number = None
    for it in iter_ldjson(html):
        if not isinstance(it, dict):
            continue
        if price is None:
            off = it.get("offers")
            if isinstance(off, dict) and off.get("price"):
                price = str(off["price"])
            elif isinstance(off, list):
                for o in off:
                    if isinstance(o, dict) and o.get("price"):
                        price = str(o["price"]); break
        # Model Number is exposed as a PropertyValue in additionalProperty
        for prop in it.get("additionalProperty", []) or []:
            if isinstance(prop, dict) and str(prop.get("name", "")).lower() == "model number":
                model_number = prop.get("value")
    # regex fallback for model number if not in a parsed object
    if not model_number:
        m = re.search(r'"name"\s*:\s*"Model Number"\s*,\s*"value"\s*:\s*"([^"]+)"', html)
        if m:
            model_number = m.group(1)

    # -------- PRICE fallback: main visible price element (scoped) --------
    if not price:
        soup = BeautifulSoup(html, "lxml")
        el = soup.select_one(
            "div[data-testid='customer-price'] span, "
            "span.font-sans.text-default.text-style-body-md-400.font-500.text-7.leading-7"
        )
        if el:
            price = el.get_text(strip=True)

    print("\n--- Extracted Product Data (BestBuy) ---")
    print(f"File: {input_file}")
    print(f"Price: {price}")
    print(f"Model Number: {model_number}")
    print("--------------------------------\n")

    return price, model_number, False

async def save_bestbuy_htmls(
    urls,
    output_dir="outputs",
    cookies_file="bestbuy_cookies.json",
    headless=True,
):
    """
    Loop over list of BestBuy URLs, save each page's HTML to output_dir,
    parse with parse_bestbuy_html and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        # --disable-http2: BestBuy's CDN breaks Playwright's HTTP/2 on the newer
        # S26 product pages (net::ERR_HTTP2_PROTOCOL_ERROR), so those pages never
        # loaded/saved. Forcing HTTP/1.1 makes them load reliably.
        browser = await p.chromium.launch(
            headless=headless, slow_mo=100, args=["--disable-http2"]
        )

        # Load existing cookies/session state if available
        if os.path.exists(cookies_file):
            print("🍪 Loading existing cookies/session...")
            context = await browser.new_context(storage_state=cookies_file)
        else:
            print("🆕 No cookies found, creating a new session...")
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 768},
            )

        results = []
        try:
            for idx, url in enumerate(urls, start=1):
                # empty slot (e.g. product not yet listed): keep the position so
                # results stay aligned with the product groups, but skip cleanly.
                if not url or not url.strip():
                    print(f"\n[BestBuy {idx}/{len(urls)}] empty URL slot -> skipping")
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": "empty"})
                    continue
                safe_name = sanitize_filename(url)[:120]
                output_file = os.path.join(output_dir, f"bestbuy_{idx}_{safe_name}.html")
                page = None
                try:
                    page = await context.new_page()
                    print(f"\n[BestBuy {idx}/{len(urls)}] Navigating to {url} ...")


                    try:

                        await page.goto(url, wait_until="load")
                        await asyncio.sleep(10)  # extra wait to ensure stability
                    except TimeoutError as te:
                        print(f"⚠️ 'load' timeout for {url} after 30s. Continuing anyway...")
                        try:
                            await page.wait_for_load_state("domcontentloaded", timeout=7000)
                        except TimeoutError:
                            pass







                    # Wait randomly for page content to settle
                    await human_delay(3, 6)

                    # 🖱️ Simulate random human-like mouse movement
                    for _ in range(3):
                        x = random.randint(200, 800)
                        y = random.randint(200, 600)
                        await page.mouse.move(x, y, steps=random.randint(5, 15))
                        await human_delay(0.3, 1.5)

                    # 🖱️ Random scrolling
                    for _ in range(2):
                        scroll_y = random.randint(400, 1000)
                        await page.mouse.wheel(0, scroll_y)
                        await human_delay(1, 3)

                    # Extract HTML (resilient to BestBuy's mid-load client-side
                    # navigation which used to make page.content() throw)
                    html_content = await get_page_content_safe(page)
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print(f"✅ HTML saved to {output_file}")

                    # Parse and collect results (updated: redirect-aware, JSON-LD)
                    price, model, redirected = parse_bestbuy_html(output_file, expected_url=url)

                    if redirected:
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "redirect"})
                    elif not price:
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "model": NOT_AVAILABLE, "status": "no_price"})
                    else:
                        results.append({"url": url, "file": output_file, "price": price, "model": model, "status": "ok"})
                    await page.close()
                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        if page:
                            await page.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"})

            # Save cookies/session state after processing all pages
            storage_state = await context.storage_state()
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(storage_state, f, ensure_ascii=False, indent=4)
            print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

# -----------------------
# SAMSUNG-specific logic
# -----------------------
async def wait_network_idle(page, timeout=15000):
    """Wait until network becomes idle (0 active requests)."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
        await asyncio.sleep(10)  # extra wait to ensure stability | not sure if this is the right approach
    except TimeoutError:
        print("⚠️ networkidle timeout — continuing anyway")

def extract_sku_from_url(url: str):
    """Extract SKU value from the given URL (looks for 'sku-<value>' or 'sm-<value>')."""
    if not url:
        return None
    
    # Try to find sku- first
    m = re.search(r"sku-([A-Za-z0-9-]+)", url, re.IGNORECASE)
    if m:
        return m.group(1)
    
    # If not found, try to find sm-
    m = re.search(r"(sm-[A-Za-z0-9-]+)", url, re.IGNORECASE)
    if m:
        return m.group(1)
    
    return None


def extract_price(filename, expected_url=None):
    """Return (price_text_or_None, redirected_bool) for a saved Samsung page.

    Updated for the current UI: the old #device_info aria-checked radios are gone.
    The reliable source is the JSON-LD Product offer, keyed to the exact SKU, so
    we never pick up a sibling variant's price. Redirects are detected via the
    page's canonical link.
    """
    if not os.path.exists(filename):
        print(f"❌ File not found for parsing: {filename}")
        return None, False

    with open(filename, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()

    expected_sku = extract_sku_from_url(expected_url) if expected_url else None
    expected_sku = expected_sku.lower() if expected_sku else None

    # -------- REDIRECT DETECTION --------
    canonical = get_canonical_href(html)
    if expected_sku and canonical:
        can_sku = extract_sku_from_url(canonical)
        can_sku = can_sku.lower() if can_sku else None
        if can_sku and can_sku != expected_sku:
            print(f"[REDIRECT] requested {expected_sku} but page is {can_sku} -> not available")
            return None, True

    # -------- PRICE from JSON-LD offer keyed to the SKU --------
    price = None
    for it in iter_ldjson(html):
        if isinstance(it, dict) and it.get("sku"):
            if expected_sku and str(it["sku"]).lower() != expected_sku:
                continue
            off = it.get("offers")
            if isinstance(off, dict) and off.get("price"):
                price = str(off["price"]); break

    print("🔎 Extracted Price:", price)
    return price, False

async def save_samsung_htmls(
    urls,
    output_dir="outputs",
    cookies_file="samsung_cookies.json",
    headless=True,
):
    """
    Loop over list of Samsung product URLs, save each page's HTML to output_dir,
    parse price and sku using the same logic you provided, and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)   # keep visible by default per original
        # Create or reuse context
        # if os.path.exists(cookies_file):
        #     print("🍪 Loading existing cookies/session...")
        #     context = await browser.new_context(storage_state=cookies_file)
        # else:
        print("🆕 No cookies found, creating a new session...")

        results = []
        try:
            for idx, url in enumerate(urls, start=1):
                # empty slot (e.g. product not yet listed): keep the position so
                # results stay aligned with the product groups, but skip cleanly.
                if not url or not url.strip():
                    print(f"\n[Samsung {idx}/{len(urls)}] empty URL slot -> skipping")
                    results.append({"url": url, "file": None, "price": None, "sku": None, "status": "empty"})
                    continue
                safe_name = sanitize_filename(url)
                output_file = os.path.join(output_dir, f"samsung_{idx}_{safe_name}.html")

                context = None
                page = None
                try:
                    # create a fresh context per-URL (ensures cookies/storage are isolated and destroyed on context.close())
                    context = await browser.new_context(
                        user_agent=(
                            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                            "AppleWebKit/537.36 (KHTML, like Gecko) "
                            "Chrome/120.0.0.0 Safari/537.36"
                        ),
                        viewport={"width": 1600, "height": 900},
                    )

                    page = await context.new_page()
                    print(f"\n[Samsung {idx}/{len(urls)}] Navigating to {url} ...")
                    await page.goto(url, wait_until="domcontentloaded")

                    print("Waiting for network to be idle...")
                    await wait_network_idle(page, timeout=20000)

                    print("Waiting for #device_info box...")
                    try:
                        await page.wait_for_selector("#device_info", timeout=20000)
                    except TimeoutError:
                        print("❌ #device_info did NOT load — Samsung blocked or loaded too slowly.")
                        # Still save HTML for debugging
                        html = await get_page_content_safe(page)
                        with open(output_file, "w", encoding="utf-8") as f:
                            f.write(html)
                        sku = extract_sku_from_url(url)
                        results.append({"url": url, "file": output_file, "price": None, "sku": sku, "status": "partial: no device_info"})
                        await page.close()
                        # close the context for this URL before continuing
                        await context.close()
                        continue

                    # Extra wait for prices inside #device_info
                    await page.wait_for_selector("#device_info span", timeout=15000)

                    # Save HTML (resilient to any mid-load client-side navigation)
                    html = await get_page_content_safe(page)
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html)
                    print(f"✅ HTML saved to {output_file}")

                    # Save cookies/session state after each page optionally (we'll write final at end too)
                    # storage = await context.storage_state()
                    # with open(cookies_file, "w", encoding="utf-8") as f:
                    #     json.dump(storage, f, indent=2)

                    # Parse saved HTML (updated: redirect-aware, JSON-LD by SKU)
                    price, redirected = extract_price(output_file, expected_url=url)
                    sku = extract_sku_from_url(url)

                    if redirected:
                        print("[REDIRECT] Samsung redirect -> not available")
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "sku": NOT_AVAILABLE, "status": "redirect"})
                    elif not price:
                        results.append({"url": url, "file": output_file, "price": NOT_AVAILABLE, "sku": NOT_AVAILABLE, "status": "no_price"})
                    else:
                        print("🔎 Final extracted values — Price:", price, "SKU:", sku)
                        results.append({"url": url, "file": output_file, "price": price, "sku": sku, "status": "ok"})

                    # tiny cooperative yield
                    await human_delay_short()

                    await page.close()
                    # close the context for this URL (destroys cookies/storage)
                    await context.close()

                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        if page:
                            await page.close()
                    except Exception:
                        pass
                    try:
                        if context:
                            await context.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "sku": None, "status": f"error: {e}"})

            # Write cookies/session state once more at the end
            # storage = await context.storage_state()
            # with open(cookies_file, "w", encoding="utf-8") as f:
            #     json.dump(storage, f, indent=2)
            # print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

# -----------------------
# Combined main
# -----------------------
async def main():
    # Replace/extend these lists with the product URLs you want to iterate over
    amazon_urls = [
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7JRKGH1/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7",
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Warranty/dp/B0DP3G4GVQ/ref=sr_1_1_sspa?crid=10TW4LFRAIOBO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7bUk0vIwEpi-0APakuwi3hHMu2vGmVltlmCoeqExLjwwHe1NY_y-eiRAZze4TELqwF9A5Z3q2WMC2EPG0p4nD5aGis4NWae_K-CRmvy0IwyOTABmJrdT_nBArRg_3HUXEeD8RiVcw9SrqiFQb-CKPztbZuf4z8k2ncgbVn8qKqGMwy7rSG9Br5vXcD_F-IobKCrdhThEoUQ0RDqrmYpPZPI.EY_aE-DTUSHzDMHcvx2u1pDmmaEgKrcYAQ31_D796hk&dib_tag=se&keywords=s25%2Bultra&qid=1763542293&sprefix=s25%2Caps%2C425&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1",
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7K3FZ79/ref=sr_1_1_sspa?crid=ZFFPIZBZ98GJ&dib=eyJ2IjoiMSJ9.UCywTCyyKG4bvq7perU6WJwDnwocjQBoU_CBTt0iLEilFUxs7eGFZYXZpU_ioObwnwWuyf6rjjxKURGHvrFykwP0YDyTNEHIJ6iMdK6L--UC4Xf9otHkBAGnuMrKXhDVPrKXBcX3EASPQMHPmIxeZyAUQDkEAC7kjvwYOc851BfCkl7yfIKNjFbb5-rq1n_ZNuEDFncqGGsmuRczfLMtzrq4HfNQyihnc2SfvotrbB8.gsaBvxdWlghc5_yM6ADh4JkaCZsCGwTYiwGF4rklYiw&dib_tag=se&keywords=galaxy%2Bz%2Bflip%2B7&qid=1763542634&sprefix=galaxy%2Bz%2Bflip7%2B%2Caps%2C427&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1",
    "https://www.amazon.com/SAMSUNG-Smartphone-Processor-ProScaler-Manufacturer/dp/B0DYVMVZSY/ref=sr_1_1_sspa?crid=18N7Z6JOQP2BV&dib=eyJ2IjoiMSJ9.nzLYfcsJ7KheFLAc8b9qkf36-GLK18wvNZAtoNJSu1Zuk0LTOxwvIqqD7blO0fqQDvGW1a_cFlDg5Nh6UJs25ksORtqvnynWCvMs3mnXvO49ZOy1Lc0OCa8xgu_zDwki3AucEZejB1tiHQzt8KYuAH3-YcGmTnO7s-Wn_1i_JPAcSstuLawUyqxRadquHocmToV-_PuNbtIeLyuTmsuGn88G4Hs_fJCfV7dzS_zI9l8.tjLmW8sTlnkCiJoCZPOQgL3qCTpiGvJ6gp9QKjm7R24&dib_tag=se&keywords=galaxy%2Bs25%2Bedge&qid=1763542733&sprefix=galaxy%2Bs25%2Bed%2Caps%2C351&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1",
    # "https://www.amazon.com/Samsung-Galaxy-America-SM-X930-SM-X930NZSAGTO/dp/B0FQKSCX2D/ref=sr_1_1?crid=3M1J0R493M49G&dib=eyJ2IjoiMSJ9.gl7sV5EmfcxLIXPuTsNfI7op0o_zOMo7gfDS2RBeRDauOS7q5VHVcvkwTTT0IZG-uVm80pmlYn8QxFoHJXj77LCR6SAqze32C5FkFT6b2LCma3NyZVHieIZPdE1gYzqsd2dpGUmcutT28tvpam9KfKaR61OBwhng8XVOKUIUAhdpovFq7bcDy2-e2HVcDWsoTl4NQBA8ox3ML2FkwpFZBFv2hbBlJWkYfN2t04lIO3E.obtlHz9d7cecPNUhBbVBtZbaggjQqd0Pw7lOv_ynn6U&dib_tag=se&keywords=Galaxy%2BTab%2BS11%2BUltra%2B14.6%22&qid=1764915926&sprefix=galaxy%2Btab%2Bs11%2Bultra%2B14.6%2B%2Caps%2C323&sr=8-1&th=1",
    "https://www.amazon.com/Samsung-Android-Display-Performance-Note-Taking/dp/B0FKLRP6MT/ref=sr_1_2?crid=IRWTTSR9PD6Z&dib=eyJ2IjoiMSJ9.gl7sV5EmfcxLIXPuTsNfI7op0o_zOMo7gfDS2RBeRDbcUoXl9R3vqob2YHTCDrovf3-tNFlHFWJY8EqEh7VuTrCR6SAqze32C5FkFT6b2LCjp2-lRnlgcI7zmK9NpQGio14ileVGjyBZ1vJGWPU_CMwActHF5VFeoaLVOkSTa2ImkBLodJGj5yfcU61SB3t8x102OAYCHSsCddgonqJQ1py6Tc_2EYjh_DQeXvcIiHM.xyJFuQBGqB1Lu5oSYjpdPwvbs9I0AaP3lfRsXFrM1Rk&dib_tag=se&keywords=Galaxy%2BTab%2BS11%2B14.6%22&qid=1764916487&sprefix=galaxy%2Btab%2Bs11%2B14.6%2B%2Caps%2C333&sr=8-2&th=1",
    # #watch ultra
    # "https://www.amazon.com/Samsung-Smartwatch-Titanium-Advanced-Coaching/dp/B0F7PKN2C8/ref=sr_1_1?crid=3SRX0N61TD6U6&dib=eyJ2IjoiMSJ9.3IBJcvYi9CChdwEbJv-M0JI20M4KL5zbOWz0RW6vcq334Dh6E6cEn86tebR0SsocNdVg9Vfz46VEQQNuG0FZvlUMYb4SLKwoQttV0nLJndbo8nzwoJThK5IXghklqyVGNAOtTAZwtnHdRVq1_GR47W9jHBRbRgbZcCOJtIlTw3bv7sBEF30rVWSOj1HzU3u9OT6ijzvnY1rwufcUjvMQg1SE2zfvWhvMQaX5TjcMIXA.pDuCRtfVtzmqmKdh98pIvWKIqi58y6aAvqv8bK-zL3w&dib_tag=se&keywords=Galaxy%2BWatch%2BUltra%2B(2025)&qid=1764917703&sprefix=galaxy%2Bwatch%2Bultra%2B2025%2B%2Caps%2C882&sr=8-1&th=1",
    # #wath 8 classic SM-L505UZKQXAA
    # "https://www.amazon.com/Samsung-Smartwatch-Rotating-Coaching-Warranty/dp/B0F7PXV6LS/ref=sr_1_2?crid=WW4NH8L6HZ5D&dib=eyJ2IjoiMSJ9.o36AO3N47s3Xupf6NuDGv61Z4rvpJOACuIl3dwe0mWJ0gGpb4SVSC3Csts97aT6rEDWw8CVBbJWBGDZBSNe_2vUS6d1wDqkljvRc7cHxEjpiJV9TlurdkZ5VwalrTrZUybfCHWR62MbXfx0_afRxif_aQPBGg8n9Bkpqdc_X5NbATMA-XgQfEQC_d4LL4KFOwPucXN93zEnBlnIfHKHwShUQmQ9GpqY5BG1g1xDGX94.tQWQuCrpym1uKokkf5YaSGbpPmRqxcmusFnsff5EPDk&dib_tag=se&keywords=Galaxy%2BWatch8%2BClassic&qid=1764917793&sprefix=galaxy%2Bwatch8%2Bclassic%2Caps%2C515&sr=8-2&th=1",
    # #BUDS 3 PRON WHITE SM-R630NZWEXAR
    # "https://www.amazon.com/SAMSUNG-Cancelling-Optimization-Interpreter-Redesigned/dp/B0D9YZ1B72/ref=sr_1_1?crid=LJ1KR3LTUZ0E&dib=eyJ2IjoiMSJ9.RbhzvfQA2eVDyW4vfE0yGXTlxYjAlt6dRnFhE49E7T6W-msgtnudbY7NKvo-7QiJqrZTjSGZ85kWCjZDydIS7l6jl4-mMdixSPjOZ69VmMbMWfiBGx6zkz0srzaWl_uuq9-TnNpml7LpJ30oYhYhEfVvEwTCBrd12UXkjZgxoODGt4NPoxTUQ74GPuf9FXT3zPdefZaGSqqFZEc2wtDXCzuTevt2p8vfx2yMKoe7aUA.pj-hN56CVzQ5a7ygsRTxXrM89X2m8BYQD6TAFnVbwCI&dib_tag=se&keywords=Galaxy%2BBuds3%2BPro&qid=1764918515&sprefix=galaxy%2Bbuds3%2Bpro%2Caps%2C538&sr=8-1&th=1",
    #other z fold7 storage sizes
    #256gb Blue Shadow
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7JZYWLL/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7&th=1",
    # 256gb Jetblack
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7K3FQN1/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7&th=1",
    #256gb Silver Shadow
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7JZXMQ7/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7&th=1",

    #1tb zet black
    "https://www.amazon.com/Samsung-Galaxy-Fold7-SM-F966U-Smartphone/dp/B0FK1XVXTB?ref_=ast_sto_dp",
    #512gb zet black
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7K9LFCL/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7&th=1",
    #512gb silver shadow
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7J243YH/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7&th=1",
    






    #s25 ultra 256gb titanium black
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3GQ4QY/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 256gb titanium gray
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3GDTCF/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 256gb titanium silver blue
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3CP2SY/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 256gb titanium white silver
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3GP92Z/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 512gb titanium gray
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP5S1MQJ/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 512gb titanium silver blue
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3F8GK1/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 512gb titanium white silver
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Silverblue/dp/B0DP3GBGCF/ref=sr_1_1?crid=2OCXY9JS2C2BO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7dPj7jSJEhMlQYfh62EDEp3xILbjXGGsUY5dJ8J_RqFmPrU9Wk-ajkTu-1OGBCqRjZOKNruE2tli1gYxD7MwgPXT7xP_Rj9bebKJyLidPy47j0YHo7dLzNITaWD7QcaCCIXMi5J1bUWvGwfQfOufUTKSOt14lWodwgk1Z5S5f_OMUvVwsk7fedeFppWKpdPUQ9Ximpz2_syOzHf193Tw0uY.AXDVzX7Lxk4fuhf3tpvgRQZNeXbGQBpVEvS8lcbL734&dib_tag=se&keywords=s25%2Bultra&qid=1768041733&sprefix=s25%2Bultr%2Caps%2C342&sr=8-1&th=1",
    #s25 ultra 1tb titanium black
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzkfxaa/",
    #s25 ultra 1tb silver blue
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzbfxaa/",
    #s25 ultra 1tb white silver
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzsfxaa/",
    #s25 edge 256gb titanium silver
    "https://www.amazon.com/SAMSUNG-Smartphone-Processor-ProScaler-Manufacturer/dp/B0DYVPJ5KD/ref=sr_1_2?crid=1M68RADQF1L03&dib=eyJ2IjoiMSJ9.WYBGTsYsYreUEDuwL6ibZ_36-GLK18wvNZAtoNJSu1YfSlow3iojC_CNZjzSXBHUfF8x3q10WG2dw8_YoqqG1RxCz_N5mIp-qXPEVsF_w0mU2XcJQ6kh45H2jGnWLGVoj2LQ-_256xDdaf224kgBrgDpK9TPvsn_opygq0fcUF-Dns9GqO_cbFVasK3cm_7JZMcIlp39FhZTYghIKBrXwFZb6MwuHvGEJf9vhG3xddc.XYKlktc015l2spoRA2oRTCbQ27oapL8ew-j_OcdUUcg&dib_tag=se&keywords=s25%2Bedge%2B256gb%2Btitanium%2Bsilver&qid=1768045803&sprefix=s25%2Bedge%2B256gb%2Btitanium%2Bsilver%2Caps%2C964&sr=8-2&th=1",
    #s25 edge 512gb titanium jet black
    "https://www.amazon.com/SAMSUNG-Smartphone-Processor-ProScaler-Manufacturer/dp/B0DYVMVZSY/ref=sr_1_2?crid=1M68RADQF1L03&dib=eyJ2IjoiMSJ9.WYBGTsYsYreUEDuwL6ibZ_36-GLK18wvNZAtoNJSu1YfSlow3iojC_CNZjzSXBHUfF8x3q10WG2dw8_YoqqG1RxCz_N5mIp-qXPEVsF_w0mU2XcJQ6kh45H2jGnWLGVoj2LQ-_256xDdaf224kgBrgDpK9TPvsn_opygq0fcUF-Dns9GqO_cbFVasK3cm_7JZMcIlp39FhZTYghIKBrXwFZb6MwuHvGEJf9vhG3xddc.XYKlktc015l2spoRA2oRTCbQ27oapL8ew-j_OcdUUcg&dib_tag=se&keywords=s25%2Bedge%2B256gb%2Btitanium%2Bsilver&qid=1768045803&sprefix=s25%2Bedge%2B256gb%2Btitanium%2Bsilver%2Caps%2C964&sr=8-2&th=1",










    #s26 ultra 256gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SX6C1G/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 256gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SX79VJ/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 256gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SX9GDX/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 256gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SWDH8P/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    #s26 ultra 512gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SVV79P/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 512gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SWN34T/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 512gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SZ6SJY/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 ultra 512gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SW3XXP/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    #s26 ultra 1tb sky blue
    "",
    #s26 ultra 1tb cobalt violet
    "",
    #s26 ultra 1tb white
    "",
    #s26 ultra 1tb black
    "",


    #s26 256gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SWZP8B/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 256gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SXD3TJ/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 256gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SYJK4G/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 256gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SW96R4/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    #s26 512gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0GH3474YH/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 512gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0GH33V5NM/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 512gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0GH3B3PFH/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26 512gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0GH33YP71/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    #s26+ 256gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SYN2QC/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 256gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SVXVSP/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 256gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SYZVKC/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 256gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SWQC2Q/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    #s26+ 512gb sky blue
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SVFSNP/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 512gb cobalt violet
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SWDNBB/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 512gb white
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SYWBQP/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",
    #s26+ 512gb black
    "https://www.amazon.com/Samsung-Unlocked-Smartphone-Charging-Warranty/dp/B0G4SSKLJM/ref=sr_1_2?crid=2UQYGUVE2NW9I&dib=eyJ2IjoiMSJ9.dmpAJBmc0kdQ-0UZY3qSO1kfTmE_zG7-B3BnDPlrHT4unck8bEo6awgs2XqYKN1Ioxsa3pqLoILDdlRDCoO2UzJ6bFYCr0F-44XU54bpcdiRCq9nQbtmhunaCfkopuNvY4QPN0zIHShJv2XmQUr_IRwyMskP7AwWQ6lH3oL1cSVZzgUSwhMskaIx2DiHwspx75klEDAQogS_AqpGWLzaCJFh6oc7DmjnWjhk3x4OrmA.Si2kDuvaGVezxppsFXjGC2L4g9-z8t0vUTmvCGDcVBI&dib_tag=se&keywords=s26%2Bultra&qid=1772457385&sprefix=s26%2Bult%2Caps%2C439&sr=8-2&th=1",


    ]

    bestbuy_urls = [
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-512gb-unlocked-blue-shadow/JJGRF3XK3P",
        "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-512gb-unlocked-titanium-black/J3ZYG25H6J",
        "https://www.bestbuy.com/product/samsung-galaxy-z-flip7-512gb-unlocked-jet-black/JJGRF335X6",
        "https://www.bestbuy.com/product/samsung-galaxy-s25-edge-512gb-unlocked-titanium-jet-black/JJGRF3CQKC",
        # Tabs
        # "https://www.bestbuy.com/product/samsung-galaxy-tab-s11-ultra-14-6-256gb-wi-fi-with-s-pen-gray/JJGRF39W23/sku/6641197",
        "https://www.bestbuy.com/product/samsung-galaxy-tab-s11-11-512gb-wi-fi-with-s-pen-gray/JJGRF39W8Q",
        # #watch ultra
        # "https://www.bestbuy.com/product/samsung-galaxy-watch-ultra-titanium-smartwatch-47mm-lte-titanium-blue-2025/JJGRF3SQXX/sku/6635999",
        # #watch 8 classic SM-L505UZKAXAA
        # "https://www.bestbuy.com/product/samsung-galaxy-watch8-classic-stainless-steel-smartwatch-46mm-lte-black-2025/JJGRF32VWX",
        # #BUDS 3 PRO SM-R630NZWAXAR
        # "https://www.bestbuy.com/product/samsung-galaxy-buds3-pro-wireless-earbud-headphones-white/J3ZYG2KR7H/sku/6585613",
        #256gb blue shadow
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-256gb-unlocked-blue-shadow/JJGRF3XK8C",
        #256gb zet black
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-256gb-unlocked-jet-black/JJGRF3XKX4",
        #256gb silver shadow
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-256gb-unlocked-silver-shadow/JJGRF3XK2Q",
        #1tb zet black
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-1tb-unlocked-jet-black/JJGRF3XK25",
        #512gb zet black
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-512gb-unlocked-jet-black/JJGRF3XKXR",
        #512gb silver shadow
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-512gb-unlocked-silver-shadow/JJGRF3XK8Y",








    #s25 ultra 256gb titanium black
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-256gb-unlocked-titanium-black/J3ZYG25JP5",
    #s25 ultra 256gb titanium gray
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-256gb-unlocked-titanium-gray/J3ZYG25HGR",
    #s25 ultra 256gb titanium silver blue
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-256gb-unlocked-titanium-silverblue/J3ZYG25H48",
    #s25 ultra 256gb titanium white silver
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-256gb-unlocked-titanium-whitesilver/J3ZYG2VPX3",
    #s25 ultra 512gb titanium gray
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-512gb-unlocked-titanium-gray/J3ZYG25TPR",
    #s25 ultra 512gb titanium silver blue
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-512gb-unlocked-titanium-silverblue/J3ZYG25HW9",
    #s25 ultra 512gb titanium white silver
    "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-512gb-unlocked-titanium-whitesilver/J3ZYG2VP76",
    #s25 ultra 1tb titanium black
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzkfxaa/",
    #s25 ultra 1tb silver blue
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzbfxaa/",
    #s25 ultra 1tb white silver
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzsfxaa/",
    #s25 edge 256gb titanium silver
    "https://www.bestbuy.com/product/samsung-galaxy-s25-edge-256gb-unlocked-titanium-silver/JJGRF3CQHC",
    #s25 edge 512gb titanium jet black
    "https://www.bestbuy.com/product/samsung-galaxy-s25-edge-512gb-unlocked-titanium-jet-black/JJGRF3CQKC",




















    #s26 ultra 256gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-256gb-unlocked-sky-blue/JJGRF36LTY",
    #s26 ultra 256gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-256gb-unlocked-cobalt-violet/JJGRF36YJC",
    #s26 ultra 256gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-256gb-unlocked-white/JJGRF36L4Q",
    #s26 ultra 256gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-256gb-unlocked-black/JJGRF36YGZ",


    #s26 ultra 512gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-512gb-unlocked-sky-blue/JJGRF36LRF",
    #s26 ultra 512gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-512gb-unlocked-cobalt-violet/JJGRF36LVC",
    #s26 ultra 512gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-512gb-unlocked-white/JJGRF36LZ2",
    #s26 ultra 512gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-512gb-unlocked-black/JJGRF36Y3Q",

    
    #s26 ultra 1tb sky blue
    "",
    #s26 ultra 1tb cobalt violet
    "",
    #s26 ultra 1tb white
    "",
    #s26 ultra 1tb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-ultra-1tb-unlocked-black/JJGRF36Y3L",


    #s26 256gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-sky-blue/JJGRF36PKF",
    #s26 256gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-cobalt-violet/JJGRF36PHF",
    #s26 256gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-white/JJGRF36PH8",
    #s26 256gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-black/JJGRF36PKH",


    #s26 512gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-sky-blue/JJGRF3TX6S",
    #s26 512gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-cobalt-violet/JJGRF3TCZR",
    #s26 512gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-white/JJGRF3TXZ4",
    #s26 512gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-black/JJGRF3T2VP",


    #s26+ 256gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-sky-blue/JJGRF36S26",
    #s26+ 256gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-cobalt-violet/JJGRF36SCG",
    #s26+ 256gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-white/JJGRF36SP3",
    #s26+ 256gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-256gb-unlocked-black/JJGRF368TV",


    #s26+ 512gb sky blue
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-sky-blue/JJGRF36PFQ",
    #s26+ 512gb cobalt violet
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-cobalt-violet/JJGRF36P52",
    #s26+ 512gb white
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-white/JJGRF36SLJ",
    #s26+ 512gb black
    "https://www.bestbuy.com/product/samsung-galaxy-s26-512gb-unlocked-black/JJGRF36SP9",











    ]

    samsung_urls = [
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-512gb-unlocked-sku-sm-f966udbexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-512gb-unlocked-sku-sm-s938uzkexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-z-flip7/buy/galaxy-z-flip7-512gb-unlocked-sku-sm-f766uzkexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-edge-512gb-unlocked-sku-sm-s937uzkexaa/",
    #tabs
    # "https://www.samsung.com/us/tablets/galaxy-tab-s11/buy/galaxy-tab-s11-ultra-256gb-gray-wi-fi-sku-sm-x930nzaaxar/",
    "https://www.samsung.com/us/tablets/galaxy-tab-s11/buy/galaxy-tab-s11-512gb-gray-wi-fi-sku-sm-x730nzaixar/",
    # #watch ultra
    # "https://www.samsung.com/us/watches/galaxy-watch-ultra-2025/buy/galaxy-watch-ultra-47mm-titanium-blue-sku-sm-l705uzb1xaa/",
    # #watch 8 classic sm-l505uzkaxaa
    # "https://www.samsung.com/us/watches/galaxy-watch8-classic/buy/galaxy-watch8-classic-46mm-black-lte-sku-sm-l505uzkaxaa/",
    # #BUDS 3 PRO sm-r630nzwaxar
    # "https://www.samsung.com/us/mobile-audio/galaxy-buds3-pro/buy/galaxy-buds3-pro-white-sm-r630nzwaxar/",
    # z fold7 other storage sizes
    #256gb blue shadow
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-256gb-unlocked-sku-sm-f966udbaxaa/",
    #256gb Jetblack
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-256gb-unlocked-sku-sm-f966uzkaxaa/",
    #256gb silver shadow
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-256gb-unlocked-sku-sm-f966uzsaxaa/",
    #1tb zetblack
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-1tb-unlocked-sku-sm-f966uzkfxaa/",
    #512gb zet black
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-512gb-unlocked-sku-sm-f966uzkexaa/",
    #512gb silver shadow
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-512gb-unlocked-sku-sm-f966uzsexaa/",






    #s25 ultra 256gb titanium black
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-256gb-unlocked-sku-sm-s938uzkaxaa/",
    #s25 ultra 256gb titanium gray
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-256gb-unlocked-sku-sm-s938uztaxaa/",
    #s25 ultra 256gb titanium silver blue
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-256gb-unlocked-sku-sm-s938uzbaxaa/",
    #s25 ultra 256gb titanium white silver
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-256gb-unlocked-sku-sm-s938uzsaxaa/",
    #s25 ultra 512gb titanium gray
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-512gb-unlocked-sku-sm-s938uztexaa/",
    #s25 ultra 512gb titanium silver blue
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-512gb-unlocked-sku-sm-s938uzbexaa/",
    #s25 ultra 512gb titanium white silver
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-512gb-unlocked-sku-sm-s938uzsexaa/",
    #s25 ultra 1tb titanium black
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzkfxaa/",
    #s25 ultra 1tb silver blue
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzbfxaa/",
    #s25 ultra 1tb white silver
    #"https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-1tb-unlocked-sku-sm-s938uzsfxaa/",
    #s25 edge 256gb titanium silver
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-edge-256gb-unlocked-sku-sm-s937uzsaxaa/",
    #s25 edge 512gb titanium jet black
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-edge-512gb-unlocked-sku-sm-s937uzkexaa/",

    



    








    #S26 Ultra
    #s26 ultra 256gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-256gb-unlocked-sku-sm-s948ulbaxaa/",
    #s26 ultra 256gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-256gb-unlocked-sku-sm-s948uzvaxaa/",
    #s26 ultra 256gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-256gb-unlocked-sku-sm-s948uzwaxaa/",
    #s26 ultra 256gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-256gb-unlocked-sku-sm-s948uzkaxaa/",


    #s26 ultra 512gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-512gb-unlocked-sku-sm-s948ulbexaa/",
    #s26 ultra 512gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-512gb-unlocked-sku-sm-s948uzvexaa/",
    #s26 ultra 512gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-512gb-unlocked-sku-sm-s948uzwexaa/",
    #s26 ultra 512gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-512gb-unlocked-sku-sm-s948uzkexaa/",



    #s26 ultra 1tb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-1tb-unlocked-sku-sm-s948ulbfxaa/",
    #s26 ultra 1tb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-1tb-unlocked-sku-sm-s948uzvfxaa/",
    #s26 ultra 1tb white
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-1tb-unlocked-sku-sm-s948uzwfxaa/",
    #s26 ultra 1tb black
    "https://www.samsung.com/us/smartphones/galaxy-s26-ultra/buy/galaxy-s26-ultra-1tb-unlocked-sku-sm-s948uzkfxaa/",





    #s26 256gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-256gb-unlocked-sku-sm-s942ulbexaa/",
    #s26 256gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-256gb-unlocked-sku-sm-s942uzvexaa/",
    #s26 256gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-256gb-unlocked-sku-sm-s942uzwexaa/",
    #s26 256gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-256gb-unlocked-sku-sm-s942uzkexaa/",


    #s26 512gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-512gb-unlocked-sku-sm-s942ulbfxaa/",
    #s26 512gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-512gb-unlocked-sku-sm-s942uzvfxaa/",
    #s26 512gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-512gb-unlocked-sku-sm-s942uzwfxaa/",
    #s26 512gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26/buy/galaxy-s26-512gb-unlocked-sku-sm-s942uzkfxaa/",


    #s26+ 256gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-256gb-unlocked-sku-sm-s947ulbaxaa/",
    #s26+ 256gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-256gb-unlocked-sku-sm-s947uzvaxaa/",
    #s26+ 256gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-256gb-unlocked-sku-sm-s947uzwaxaa/",
    #s26+ 256gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-256gb-unlocked-sku-sm-s947uzkaxaa/",


    #s26+ 512gb sky blue
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-512gb-unlocked-sku-sm-s947ulbexaa/",
    #s26+ 512gb Cobalt violet
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-512gb-unlocked-sku-sm-s947uzvexaa/",
    #s26+ 512gb white
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-512gb-unlocked-sku-sm-s947uzwexaa/",
    #s26+ 512gb black
    "https://www.samsung.com/us/smartphones/galaxy-s26-plus/buy/galaxy-s26-plus-512gb-unlocked-sku-sm-s947uzkexaa/",


    ]

    print("\n=== Running Amazon scraper ===")
    am_res = await save_amazon_htmls(amazon_urls, output_dir="outputs", cookies_file="amazon_cookies.json", headless=True)
    print("\nAmazon Summary:")
    for r in am_res:
        print(r)

    print("\n=== Running BestBuy scraper ===")
    bb_res = await save_bestbuy_htmls(bestbuy_urls, output_dir="outputs", cookies_file="bestbuy_cookies.json", headless=True)
    print("\nBestBuy Summary:")
    for r in bb_res:
        print(r)

    print("\n=== Running Samsung scraper ===")
    sam_res = await save_samsung_htmls(samsung_urls, output_dir="outputs", cookies_file="samsung_cookies.json", headless=True)
    print("\nSamsung Summary:")
    for r in sam_res:
        print(r)

    # -----------------------
    # Write results in the SAME layout as Price Comparisons_v3_WIP.xlsx
    # (one row per run; prices + SKU columns only; 'vs' formulas left blank).
    # am_res / bb_res / sam_res are in URL order, i.e. slot order, so index s
    # maps directly to product group s.
    # -----------------------
    utc_now = datetime.datetime.now(datetime.timezone.utc)
    est_now = utc_now.astimezone(ZoneInfo("US/Eastern"))
    ts_str = est_now.strftime("%d %b %Y, %H:%M")   # e.g. "05 Dec 2025, 07:25"

    excel_file = os.path.join("outputs", "results.xlsx")
    save_results_wip_format(am_res, bb_res, sam_res, samsung_urls, ts_str, excel_file)

if __name__ == "__main__":
    asyncio.run(main())
    file_path = "outputs/results.xlsx"

    # column_references = [
    #     "a","d","ar","x","e","as","y","blank column",
    #     "i","aw","ac","j","ax","ad","blank column",
    #     "n","bb","ah","o","bc","ai","blank column",
    #     "s","bg","am","t","bh","an"
    # ]

    # created = copy_columns_by_references(
    #     file_path=file_path,
    #     column_refs=column_references,
    #     source_sheet_name=None,      # None => use first sheet; or set "Sheet1"
    #     new_sheet_base_name="SelectedColumns"
    # )
    # print(f"Created sheet: {created} in {file_path}")
