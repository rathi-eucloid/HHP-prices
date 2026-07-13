#!/usr/bin/env python3
# =========================================================================
# send_mail.py
#
# Builds a professional HTML price-comparison report from outputs/results.xlsx
# and emails it (inline + as an .html attachment) through Microsoft's SMTP.
#
# Layout mirrors script_3.py's results.xlsx (see PRODUCT groups below) and the
# formatting reference in image.png:
#   * one shared "Timestamp EST" column on the left
#   * each product is a group of columns with a merged product title on top
#   * per product we ONLY show:  SKU | Samsung Price | Amazon Price | vs Amazon
#       - Samsung price column first, then Amazon price column
#       - BestBuy price / BestBuy & Amazon SKU columns are dropped
#       - only ONE sku column (Samsung's), header simply reads "SKU"
#       - "vs Amazon" is shown as a percentage (e.g. +3.5% / -13.5%),
#         NOT the raw -1.00..1.00 delta
#   * a narrow blank "column breaker" separates one product from the next
#   * only rows from the past 15 days (relative to send time) are included
#
# Credentials & recipients come from environment variables (never hard-coded):
#   MAIL_USERNAME   sender mailbox / login          (required)
#   MAIL_PASSWORD   mailbox password / app password (required)
#   MAIL_TO         recipient(s), comma-separated    (required)
#   MAIL_FROM       From address (defaults to MAIL_USERNAME)
#   MAIL_SUBJECT    subject line   (optional, sensible default)
#   SMTP_HOST       defaults to smtp.office365.com   (Microsoft default)
#   SMTP_PORT       defaults to 587 (STARTTLS)
# =========================================================================
import os
import ssl
import sys
import smtplib
from datetime import datetime, timedelta
from email.message import EmailMessage
from email.utils import formatdate, make_msgid

from openpyxl import load_workbook

# ---- must match script_3.py's WIP layout ------------------------------------
EXCEL_PATH      = os.path.join("outputs", "results.xlsx")
FIRST_GROUP_COL = 3          # column C
GROUP_STRIDE    = 9
TIMESTAMP_COL   = 2          # column B
HEADER_ROW      = 1          # product labels
SUBHEADER_ROW   = 2          # sub-headers
FIRST_DATA_ROW  = 3
TS_FORMAT       = "%d %b %Y, %H:%M"   # e.g. "05 Jul 2026, 08:02"
NOT_AVAILABLE   = "not available"
WINDOW_DAYS     = 15

# per-group column offsets (0-indexed from the group's first column)
OFF_AMAZON_PRICE  = 0
OFF_SAMSUNG_PRICE = 1
OFF_SKU_SAMSUNG   = 4

# ---- Microsoft SMTP defaults ------------------------------------------------
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.office365.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _as_float(value):
    """Return a float for a real price cell, or None for blanks/'not available'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "" or s.lower() == NOT_AVAILABLE:
        return None
    try:
        return float(s.replace(",", "").replace("$", ""))
    except ValueError:
        return None


_NOT_FETCHED = ('<span style="font-size:11px;font-weight:400;'
                'font-style:italic;color:#9e9e9e;">data not fetched</span>')


def _fmt_price(value):
    f = _as_float(value)
    if f is None:
        return _NOT_FETCHED
    return f"${f:,.2f}"


def _fmt_sku(value):
    if value is None:
        return _NOT_FETCHED
    s = str(value).strip()
    if s == "" or s.lower() == NOT_AVAILABLE:
        return _NOT_FETCHED
    return s.upper()


def _fmt_vs(amazon_val, samsung_val):
    """vs Amazon = Amazon / Samsung - 1, rendered as a signed percentage.
    Returns (text, colour). Sign and colour are swapped: Amazon pricier than
    Samsung -> shown as negative -> green; Amazon cheaper -> positive -> red."""
    a = _as_float(amazon_val)
    s = _as_float(samsung_val)
    if a is None or s is None or s == 0:
        return "&mdash;", "#9e9e9e"
    pct = (a / s - 1.0) * 100.0
    if round(pct, 1) == 0:        # rounds to zero -> show a clean "0%"
        return "0%", "#e0e0e0"
    if pct > 0:
        colour = "#4caf50"        # Amazon pricier than Samsung
    else:
        colour = "#ff6b6b"        # Amazon cheaper than Samsung
    return f"{-pct:+.1f}%", colour


def _parse_ts(value):
    """Parse the timestamp cell into a datetime, or None if it can't be read."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, TS_FORMAT)
    except ValueError:
        # tolerate minor variants (e.g. a stray seconds component)
        for fmt in ("%d %b %Y, %H:%M:%S", "%d %b %Y %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


# ---------------------------------------------------------------------------
# read workbook -> in-memory model
# ---------------------------------------------------------------------------
def load_report_model(path=EXCEL_PATH):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Report source not found: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    # discover product groups from row 1 (label sits at each group's first column)
    groups = []
    s = 0
    while True:
        gc = FIRST_GROUP_COL + GROUP_STRIDE * s
        if gc > max_col:
            break
        label = ws.cell(row=HEADER_ROW, column=gc).value
        if label is not None and str(label).strip() != "":
            groups.append({"col": gc, "label": str(label).strip()})
        s += 1

    # show products in reverse order (last product first)
    groups.reverse()

    # cutoff: keep rows whose timestamp is within the past WINDOW_DAYS
    cutoff = datetime.now() - timedelta(days=WINDOW_DAYS)

    rows = []
    for r in range(FIRST_DATA_ROW, max_row + 1):
        ts_val = ws.cell(row=r, column=TIMESTAMP_COL).value
        ts = _parse_ts(ts_val)
        if ts is None or ts < cutoff:
            continue

        cells = []
        for g in groups:
            gc = g["col"]
            amazon  = ws.cell(row=r, column=gc + OFF_AMAZON_PRICE).value
            samsung = ws.cell(row=r, column=gc + OFF_SAMSUNG_PRICE).value
            sku     = ws.cell(row=r, column=gc + OFF_SKU_SAMSUNG).value
            vs_txt, vs_colour = _fmt_vs(amazon, samsung)
            cells.append({
                "sku":     _fmt_sku(sku),
                "samsung": _fmt_price(samsung),
                "amazon":  _fmt_price(amazon),
                "vs":      vs_txt,
                "vs_colour": vs_colour,
            })
        rows.append({"timestamp": str(ts_val).strip(), "cells": cells, "_ts": ts})

    # newest first: sort rows by timestamp descending
    rows.sort(key=lambda row: row["_ts"], reverse=True)

    return groups, rows, cutoff


# ---------------------------------------------------------------------------
# HTML rendering (inline styles only -- email clients strip <style> blocks)
# ---------------------------------------------------------------------------
# palette modelled on image.png (dark theme, blue accent on the "vs" column)
BG          = "#1e1e1e"
GRID        = "#000000"
TITLE_BG    = "#2b2b2b"
SUBHEAD_BG  = "#333333"
CELL_BG     = "#262626"
CELL_BG_ALT = "#2d2d2d"
TEXT        = "#f5f5f5"
ACCENT      = "#5b9bd5"     # "vs Amazon" header colour

TD_BASE = (f"border:1px solid {GRID};padding:8px 12px;"
           f"font-family:Segoe UI,Arial,sans-serif;font-size:13px;"
           f"color:{TEXT};white-space:nowrap;")


def _th(text, *, bg, colour=TEXT, colspan=1, align="center", size="13px", bold=True):
    weight = "700" if bold else "400"
    span = f' colspan="{colspan}"' if colspan > 1 else ""
    return (f'<th{span} style="{TD_BASE}background:{bg};color:{colour};'
            f'text-align:{align};font-weight:{weight};font-size:{size};">{text}</th>')


def build_table_html(groups, rows):
    # One narrow table per product, stacked vertically so the reader scrolls
    # down (not sideways). A separator sits between consecutive products.
    parts = []
    for gi, g in enumerate(groups):
        parts.append(
            f'<table cellspacing="0" cellpadding="0" '
            f'style="border-collapse:collapse;background:{BG};'
            f'border:1px solid {GRID};width:100%;max-width:640px;'
            f'margin:0 0 4px 0;">'
        )

        # ---- product title (spans all 5 columns) ----------------------------
        parts.append("<tr>")
        parts.append(_th(g["label"], bg=TITLE_BG, colspan=5, size="14px",
                         align="left"))
        parts.append("</tr>")

        # ---- sub-headers ----------------------------------------------------
        parts.append("<tr>")
        parts.append(_th("Timestamp EST", bg=SUBHEAD_BG))
        parts.append(_th("SKU", bg=SUBHEAD_BG))
        parts.append(_th("Samsung Price", bg=SUBHEAD_BG))
        parts.append(_th("Amazon Price", bg=SUBHEAD_BG))
        parts.append(_th("vs Amazon", bg=SUBHEAD_BG, colour=ACCENT))
        parts.append("</tr>")

        # ---- data rows (one per timestamp) ----------------------------------
        for ridx, row in enumerate(rows):
            cell = row["cells"][gi]
            row_bg = CELL_BG if ridx % 2 == 0 else CELL_BG_ALT
            parts.append("<tr>")
            parts.append(
                f'<td style="{TD_BASE}background:{SUBHEAD_BG};text-align:center;'
                f'font-weight:700;">{row["timestamp"]}</td>'
            )
            parts.append(
                f'<td style="{TD_BASE}background:{row_bg};text-align:center;'
                f'font-weight:600;">{cell["sku"]}</td>'
            )
            parts.append(
                f'<td style="{TD_BASE}background:{row_bg};text-align:right;'
                f'font-weight:700;">{cell["samsung"]}</td>'
            )
            parts.append(
                f'<td style="{TD_BASE}background:{row_bg};text-align:right;">'
                f'{cell["amazon"]}</td>'
            )
            parts.append(
                f'<td style="{TD_BASE}background:{row_bg};text-align:right;'
                f'font-weight:700;color:{cell["vs_colour"]};">{cell["vs"]}</td>'
            )
            parts.append("</tr>")

        parts.append("</table>")

        # ---- separator between products (not after the last one) ------------
        if gi != len(groups) - 1:
            parts.append(
                f'<div style="height:1px;background:{GRID};'
                f'max-width:640px;margin:18px 0;"></div>'
            )

    return "".join(parts)


def build_email_html(groups, rows, cutoff):
    if rows:
        table_html = build_table_html(groups, rows)
        note = (f'Showing {len(rows)} snapshot(s) across {len(groups)} product(s) '
                f'from the past {WINDOW_DAYS} days.')
    else:
        table_html = ('<p style="font-family:Segoe UI,Arial,sans-serif;color:#333;">'
                      f'No price snapshots were recorded in the past {WINDOW_DAYS} days.</p>')
        note = f'No data available for the past {WINDOW_DAYS} days.'

    body = f"""\
<!DOCTYPE html>
<html>
  <head><meta charset="utf-8"></head>
  <body style="margin:0;padding:24px;background:#f4f4f4;
               font-family:Segoe UI,Arial,sans-serif;color:#222;">
    <p style="font-size:15px;margin:0 0 6px 0;">Hello,</p>
    <p style="font-size:15px;margin:0 0 16px 0;">
      Please find below the latest Samsung vs Amazon price comparison report.
      The full report is also attached as an HTML file for your records.
    </p>
    <p style="font-size:12px;color:#666;margin:0 0 18px 0;">{note}</p>
    <div style="overflow-x:auto;">
      {table_html}
    </div>
    <p style="font-size:13px;color:#444;margin:18px 0 0 0;">
      Best regards,<br>Price Tracking Automation
    </p>
  </body>
</html>"""
    return body


# ---------------------------------------------------------------------------
# email sending
# ---------------------------------------------------------------------------
def send_email(html_body, attachment_html):
    username = os.environ.get("MAIL_USERNAME")
    password = os.environ.get("MAIL_PASSWORD")
    recipients_raw = os.environ.get("MAIL_TO", "")
    sender = os.environ.get("MAIL_FROM", username)
    subject = os.environ.get(
        "MAIL_SUBJECT",
        f"Samsung vs Amazon Price Report — {datetime.now():%d %b %Y}",
    )

    missing = [name for name, val in
               (("MAIL_USERNAME", username),
                ("MAIL_PASSWORD", password),
                ("MAIL_TO", recipients_raw)) if not val]
    if missing:
        raise SystemExit(f"❌ Missing required environment variable(s): {', '.join(missing)}")

    recipients = [r.strip() for r in recipients_raw.split(",") if r.strip()]

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid()

    msg.set_content(
        "Hello,\n\nPlease find attached the Samsung vs Amazon price comparison "
        "report. This message is best viewed in an HTML-capable email client.\n\n"
        "Best regards,\nPrice Tracking Automation"
    )
    msg.add_alternative(html_body, subtype="html")

    msg.add_attachment(
        attachment_html.encode("utf-8"),
        maintype="text",
        subtype="html",
        filename=f"price_report_{datetime.now():%Y-%m-%d}.html",
    )

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as server:
        server.ehlo()
        server.starttls(context=context)
        server.ehlo()
        server.login(username, password)
        server.send_message(msg, from_addr=sender, to_addrs=recipients)

    print(f"✅ Report emailed to {len(recipients)} recipient(s) via {SMTP_HOST}:{SMTP_PORT}")


def main():
    groups, rows, cutoff = load_report_model()
    print(f"Loaded {len(groups)} product group(s); "
          f"{len(rows)} row(s) within the past {WINDOW_DAYS} days.")
    email_html = build_email_html(groups, rows, cutoff)
    send_email(email_html, email_html)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # surface a clean failure in the CI job log
        print(f"❌ send_mail.py failed: {exc}", file=sys.stderr)
        raise
